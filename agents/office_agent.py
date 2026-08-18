import os
import re
import json
import requests
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# JARVIS OFFICE AGENT
# ============================================================

OLLAMA_URL = "http://localhost:11434/api/generate"
OLLAMA_MODEL = "llama3.2:3b"

BASE_DIR = os.path.abspath(
    os.path.join(
        os.path.dirname(__file__),
        "..",
    )
)

REPORTS_DIR = os.path.join(
    BASE_DIR,
    "reports",
)

os.makedirs(
    REPORTS_DIR,
    exist_ok=True,
)


# ============================================================
# OFFICE AGENT
# ============================================================

class OfficeAgent:

    def __init__(self):

        self.name = "office"

        self.reports_dir = REPORTS_DIR


    # ========================================================
    # OLLAMA
    # ========================================================

    def ask_ai(self, query):

        prompt = f"""
You are the JARVIS Office Agent.

Analyze the user's request and return ONLY valid JSON.

Do not use markdown.

Return exactly this structure:

{{
    "report_type": "sales",
    "title": "Sales Report",
    "headers": ["Date", "Product", "Quantity", "Revenue"],
    "sample_rows": [
        ["2026-01-01", "Example Product", 10, 1000]
    ]
}}

Rules:

1. report_type should describe the report.
2. title should be a useful report title.
3. headers must be a JSON array.
4. sample_rows must be a JSON array of rows.
5. If the user does not provide actual data, create a useful EMPTY TEMPLATE.
6. Do not invent real-world data.
7. Example/template data must be clearly generic.
8. Do not claim that an external file or database was accessed.
9. Do not invent tools.

USER REQUEST:

{query}
"""

        try:

            response = requests.post(
                OLLAMA_URL,
                json={
                    "model": OLLAMA_MODEL,
                    "prompt": prompt,
                    "stream": False,
                    "format": "json",
                },
                timeout=120,
            )

            response.raise_for_status()

            data = response.json()

            raw = data.get(
                "response",
                "",
            ).strip()

            return self.extract_json(
                raw
            )

        except Exception as e:

            print(
                f"\nJARVIS OFFICE AI DEBUG > {e}"
            )

            return None


    # ========================================================
    # JSON EXTRACTION
    # ========================================================

    def extract_json(self, raw):

        if not raw:

            return None

        raw = str(raw).strip()

        raw = re.sub(
            r"^```(?:json)?\s*",
            "",
            raw,
            flags=re.IGNORECASE,
        )

        raw = re.sub(
            r"\s*```$",
            "",
            raw,
        )

        try:

            data = json.loads(
                raw
            )

            if isinstance(
                data,
                dict,
            ):

                return data

        except Exception:

            pass


        start = raw.find("{")
        end = raw.rfind("}")


        if (
            start != -1
            and end != -1
            and end > start
        ):

            try:

                data = json.loads(
                    raw[
                        start:end + 1
                    ]
                )

                if isinstance(
                    data,
                    dict,
                ):

                    return data

            except Exception:

                pass


        return None


    # ========================================================
    # SAFE FILENAME
    # ========================================================

    def safe_filename(self, name):

        if not name:

            name = "office_report"

        name = str(
            name
        ).strip()

        name = re.sub(
            r"[<>:\"/\\|?*]",
            "",
            name,
        )

        name = re.sub(
            r"\s+",
            "_",
            name,
        )

        name = name.strip(
            "._ "
        )

        if not name:

            name = "office_report"

        return name[:100]


    # ========================================================
    # UNIQUE FILE PATH
    # ========================================================

    def unique_report_path(
        self,
        title,
    ):

        filename = (
            self.safe_filename(
                title
            )
            + ".xlsx"
        )

        path = os.path.join(
            self.reports_dir,
            filename,
        )

        counter = 2

        while os.path.exists(
            path
        ):

            filename = (
                self.safe_filename(
                    title
                )
                + f"_{counter}.xlsx"
            )

            path = os.path.join(
                self.reports_dir,
                filename,
            )

            counter += 1

        return path


    # ========================================================
    # CREATE EXCEL
    # ========================================================

    def create_excel_report(
        self,
        title,
        headers,
        rows,
    ):

        if not title:

            title = "JARVIS Report"


        if not isinstance(
            headers,
            list,
        ) or not headers:

            headers = [
                "Item",
                "Description",
                "Value",
            ]


        if not isinstance(
            rows,
            list,
        ):

            rows = []


        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Report"


        # ----------------------------------------------------
        # TITLE
        # ----------------------------------------------------

        worksheet["A1"] = title

        worksheet["A1"].font = Font(
            bold=True,
            size=16,
        )

        worksheet["A1"].alignment = Alignment(
            horizontal="center"
        )


        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=max(
                1,
                len(headers),
            ),
        )


        # ----------------------------------------------------
        # GENERATED DATE
        # ----------------------------------------------------

        worksheet["A2"] = (
            "Generated by JARVIS"
        )

        worksheet["A3"] = (
            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )
        )


        # ----------------------------------------------------
        # HEADERS
        # ----------------------------------------------------

        header_row = 5

        for column_index, header in enumerate(
            headers,
            start=1,
        ):

            cell = worksheet.cell(
                row=header_row,
                column=column_index,
                value=str(header),
            )

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )


        # ----------------------------------------------------
        # DATA
        # ----------------------------------------------------

        for row_index, row in enumerate(
            rows,
            start=header_row + 1,
        ):

            if not isinstance(
                row,
                list,
            ):

                continue


            for column_index, value in enumerate(
                row,
                start=1,
            ):

                worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )


        # ----------------------------------------------------
        # COLUMN WIDTH
        # ----------------------------------------------------

        for column_index in range(
            1,
            len(headers) + 1,
        ):

            letter = get_column_letter(
                column_index
            )

            maximum_length = 0


            for cell in worksheet[
                letter
            ]:

                if cell.value is None:

                    continue

                length = len(
                    str(
                        cell.value
                    )
                )

                if length > maximum_length:

                    maximum_length = length


            worksheet.column_dimensions[
                letter
            ].width = min(
                max(
                    maximum_length + 2,
                    12,
                ),
                40,
            )


        # ----------------------------------------------------
        # FREEZE HEADER
        # ----------------------------------------------------

        worksheet.freeze_panes = "A6"


        # ----------------------------------------------------
        # FILTER
        # ----------------------------------------------------

        if headers:

            last_column = get_column_letter(
                len(headers)
            )

            worksheet.auto_filter.ref = (
                f"A5:{last_column}"
                f"{max(5, header_row + len(rows))}"
            )


        # ----------------------------------------------------
        # SAVE
        # ----------------------------------------------------

        path = self.unique_report_path(
            title
        )

        workbook.save(
            path
        )

        return path


    # ========================================================
    # OFFICE REQUEST
    # ========================================================

    def office(self, query):

        print(
            "\nJARVIS OFFICE AGENT > "
            "Analyzing request"
        )


        data = self.ask_ai(
            query
        )


        # ----------------------------------------------------
        # AI FAILURE
        # ----------------------------------------------------

        if not data:

            data = {
                "report_type": "general",
                "title": "JARVIS Report",
                "headers": [
                    "Item",
                    "Description",
                    "Value",
                ],
                "sample_rows": [],
            }


        title = data.get(
            "title",
            "JARVIS Report",
        )


        headers = data.get(
            "headers",
            [],
        )


        rows = data.get(
            "sample_rows",
            [],
        )


        # ----------------------------------------------------
        # VALIDATE HEADERS
        # ----------------------------------------------------

        if not isinstance(
            headers,
            list,
        ) or not headers:

            headers = [
                "Item",
                "Description",
                "Value",
            ]


        # ----------------------------------------------------
        # VALIDATE ROWS
        # ----------------------------------------------------

        if not isinstance(
            rows,
            list,
        ):

            rows = []


        print(
            "\nJARVIS OFFICE AGENT > "
            "Creating Excel report"
        )


        try:

            path = self.create_excel_report(
                title=title,
                headers=headers,
                rows=rows,
            )


            relative_path = os.path.relpath(
                path,
                BASE_DIR,
            )


            return {
                "success": True,
                "type": "office",
                "message": (
                    "Excel report created successfully.\n"
                    f"File: {path}\n"
                    f"Report type: "
                    f"{data.get('report_type', 'general')}\n"
                    "The report is ready to open."
                ),
                "path": path,
            }


        except Exception as e:

            return {
                "success": False,
                "type": "office",
                "message": (
                    "I couldn't create the Excel report: "
                    f"{e}"
                ),
            }


# ============================================================
# GLOBAL AGENT
# ============================================================

office_agent = OfficeAgent()


# ============================================================
# PUBLIC FUNCTION
# ============================================================

def office(query):

    return office_agent.office(
        query
    )


# ============================================================
# TEST
# ============================================================

if __name__ == "__main__":

    print("=" * 60)
    print("JARVIS OFFICE AGENT TEST")
    print("=" * 60)


    result = office(
        "make an Excel sales report"
    )


    print()


    print(
        result.get(
            "message",
            result,
        )
    )