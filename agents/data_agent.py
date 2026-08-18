# ============================================================
# JARVIS DATA AGENT
# FINAL V2.5
# ============================================================

from __future__ import annotations

import os
import re
import json
import math
import warnings
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd


# ============================================================
# CONFIGURATION
# ============================================================

SUPPORTED_EXTENSIONS = {
    ".csv",
    ".xlsx",
    ".xls",
    ".json",
    ".parquet",
}

DEFAULT_MAX_ROWS = 100_000
MAX_REPORT_ROWS = 100_000
MAX_FILES_TO_DISPLAY = 30

OUTPUT_DIR = Path(
    os.environ.get(
        "JARVIS_DATA_OUTPUT",
        str(
            Path.home()
            / "Documents"
            / "JARVIS_Data_Reports"
        ),
    )
)

IGNORED_PREFIXES = (
    "~$",
    ".~",
)

IGNORED_DIRECTORIES = {
    ".git",
    ".venv",
    "venv",
    "node_modules",
    "__pycache__",
    ".git",
    ".idea",
    ".vscode",
}


# ============================================================
# OPTIONAL DEPENDENCIES
# ============================================================

try:
    import openpyxl

    OPENPYXL_AVAILABLE = True

except Exception:
    openpyxl = None
    OPENPYXL_AVAILABLE = False


try:
    import plotly.express as px

    PLOTLY_AVAILABLE = True

except Exception:
    px = None
    PLOTLY_AVAILABLE = False


# ============================================================
# PENDING DATASET SELECTION
# ============================================================

_PENDING_DATASETS: List[Path] = []
_PENDING_REQUEST = ""


def has_pending_selection():
    return bool(_PENDING_DATASETS)


def clear_pending_selection():

    global _PENDING_DATASETS
    global _PENDING_REQUEST

    _PENDING_DATASETS = []
    _PENDING_REQUEST = ""


# ============================================================
# LOGGING
# ============================================================

def log(message):

    print(
        f"\nJARVIS DATA AGENT > {message}"
    )


def debug(message):

    print(
        f"JARVIS DATA AGENT DEBUG > {message}"
    )


# ============================================================
# BASIC HELPERS
# ============================================================

def normalize(value):

    if value is None:
        return ""

    value = str(
        value
    ).strip().lower()

    value = re.sub(
        r"[^a-z0-9]+",
        " ",
        value,
    )

    return " ".join(
        value.split()
    )


def safe_string(value):

    if value is None:
        return ""

    try:

        if pd.isna(value):
            return ""

    except Exception:
        pass

    return str(value)


def json_safe(value):

    if value is None:
        return None

    if isinstance(
        value,
        (
            str,
            int,
            bool,
        ),
    ):
        return value

    if isinstance(
        value,
        float,
    ):

        if math.isnan(value):
            return None

        if math.isinf(value):
            return None

        return value

    if isinstance(
        value,
        datetime,
    ):

        return value.isoformat()

    try:

        if hasattr(
            value,
            "item",
        ):

            return json_safe(
                value.item()
            )

    except Exception:
        pass

    return str(value)


def human_size(size):

    units = [
        "B",
        "KB",
        "MB",
        "GB",
        "TB",
    ]

    value = float(size)

    for unit in units:

        if value < 1024:

            return (
                f"{value:.1f} {unit}"
            )

        value /= 1024

    return (
        f"{value:.1f} PB"
    )


def timestamp():

    return datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )


def ensure_output_dir():

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    return OUTPUT_DIR


# ============================================================
# REQUEST TYPE
# ============================================================

def requested_formats(request):

    value = normalize(
        request
    )

    return {
        "excel":
            (
                "excel" in value
                or "xlsx" in value
                or "spreadsheet" in value
            ),

        "html":
            (
                "html" in value
                or "web dashboard" in value
                or "webpage" in value
            ),

        "dashboard":
            "dashboard" in value,

    }


# ============================================================
# PATH EXTRACTION
# ============================================================

def extract_windows_path(text):

    if not text:
        return None

    text = str(text)


    # --------------------------------------------------------
    # Quoted path
    # --------------------------------------------------------

    patterns = [
        r'"([A-Za-z]:\\[^"]+)"',
        r"'([A-Za-z]:\\[^']+)'",
    ]

    for pattern in patterns:

        match = re.search(
            pattern,
            text,
            re.IGNORECASE,
        )

        if match:

            return Path(
                match.group(1)
                .strip()
                .rstrip(
                    " .,;:)"
                )
            )


    # --------------------------------------------------------
    # Known dataset file extension
    # --------------------------------------------------------

    match = re.search(

        r'([A-Za-z]:\\[^<>:"|?*\r\n]+?\.(?:csv|xlsx|xls|json|parquet))',

        text,

        re.IGNORECASE,

    )

    if match:

        return Path(
            match.group(1)
            .strip()
            .rstrip(
                " .,;:)"
            )
        )


    # --------------------------------------------------------
    # Generic Windows folder path
    # --------------------------------------------------------

    match = re.search(
        r'([A-Za-z]:\\[^\r\n]+)',
        text,
        re.IGNORECASE,
    )

    if not match:

        return None


    candidate = (
        match.group(1)
        .strip()
    )


    # Cut off natural-language instructions
    # after the actual path.

    markers = [
        " and create ",
        " and make ",
        " and generate ",
        " with ",
        " then ",
        " and show ",
        " and analyze ",
        " and analyse ",
    ]

    lower = candidate.lower()

    for marker in markers:

        index = lower.find(
            marker
        )

        if index != -1:

            candidate = candidate[
                :index
            ]

            break


    candidate = candidate.rstrip(
        " .,;:)"
    )


    return Path(
        candidate
    )


# ============================================================
# DATASET DETECTION
# ============================================================

def is_supported_file(path):

    if not path.is_file():
        return False

    if path.name.startswith(
        IGNORED_PREFIXES
    ):
        return False

    return (
        path.suffix.lower()
        in SUPPORTED_EXTENSIONS
    )


def file_type(path):

    mapping = {
        ".csv": "CSV",
        ".xlsx": "Excel XLSX",
        ".xls": "Excel XLS",
        ".json": "JSON",
        ".parquet": "Parquet",
    }

    return mapping.get(
        path.suffix.lower(),
        path.suffix.upper(),
    )


def discover_datasets(
    folder,
    recursive=True,
):

    if not folder.exists():
        return []

    if not folder.is_dir():
        return []

    datasets = []

    try:

        iterator = (
            folder.rglob("*")
            if recursive
            else folder.iterdir()
        )

        for path in iterator:

            if not path.is_file():
                continue

            if any(
                part in IGNORED_DIRECTORIES
                for part in path.parts
            ):
                continue

            if is_supported_file(path):

                datasets.append(
                    path
                )

    except Exception as e:

        debug(
            f"Dataset discovery failed: {e}"
        )


    datasets.sort(
        key=lambda path:
            path.stat().st_mtime,
        reverse=True,
    )

    return datasets


# ============================================================
# EXCEL SHEETS
# ============================================================

def get_excel_sheets(path):

    if not OPENPYXL_AVAILABLE:
        return []

    try:

        workbook = openpyxl.load_workbook(
            filename=str(path),
            read_only=True,
            data_only=True,
        )

        sheets = list(
            workbook.sheetnames
        )

        workbook.close()

        return sheets

    except Exception as e:

        debug(
            f"Excel sheet discovery failed: {e}"
        )

        return []


# ============================================================
# LOADERS
# ============================================================

def load_csv(
    path,
    max_rows,
):

    try:

        return pd.read_csv(
            path,
            nrows=max_rows,
            low_memory=False,
        )

    except UnicodeDecodeError:

        return pd.read_csv(
            path,
            nrows=max_rows,
            encoding="latin1",
            low_memory=False,
        )


def load_json(
    path,
    max_rows,
):

    with open(
        path,
        "r",
        encoding="utf-8",
    ) as file:

        data = json.load(
            file
        )


    if isinstance(
        data,
        list,
    ):

        return pd.DataFrame(
            data[:max_rows]
        )


    if isinstance(
        data,
        dict,
    ):

        for key in (
            "data",
            "records",
            "rows",
            "items",
            "results",
        ):

            value = data.get(
                key
            )

            if isinstance(
                value,
                list,
            ):

                return pd.DataFrame(
                    value[:max_rows]
                )


        return pd.DataFrame(
            [data]
        )


    return pd.DataFrame()


def load_parquet(
    path,
    max_rows,
):

    df = pd.read_parquet(
        path
    )

    if len(df) > max_rows:

        df = df.head(
            max_rows
        )

    return df


def load_excel(
    path,
    sheet_name=None,
    max_rows=DEFAULT_MAX_ROWS,
):

    sheets = get_excel_sheets(
        path
    )


    if sheets:

        if sheet_name is None:

            sheet_name = sheets[0]

        elif sheet_name not in sheets:

            try:

                index = int(
                    sheet_name
                )

                if (
                    0 <= index
                    < len(sheets)
                ):

                    sheet_name = sheets[
                        index
                    ]

                else:

                    sheet_name = sheets[0]

            except Exception:

                sheet_name = sheets[0]


    else:

        sheet_name = (
            sheet_name
            if sheet_name is not None
            else 0
        )


    try:

        with warnings.catch_warnings():

            warnings.simplefilter(
                "ignore"
            )

            return pd.read_excel(
                path,
                sheet_name=sheet_name,
                nrows=max_rows,
            )

    except Exception as e:

        raise RuntimeError(
            f"Could not read Excel sheet "
            f"'{sheet_name}' from "
            f"'{path.name}': {e}"
        ) from e


def load_dataset(
    path,
    sheet_name=None,
):

    if not path.exists():

        raise FileNotFoundError(
            str(path)
        )


    extension = (
        path.suffix.lower()
    )


    metadata = {

        "path":
            str(path),

        "name":
            path.name,

        "type":
            file_type(path),

        "size":
            human_size(
                path.stat().st_size
            ),

        "sheet":
            sheet_name,

    }


    if extension == ".csv":

        df = load_csv(
            path,
            DEFAULT_MAX_ROWS,
        )


    elif extension in {
        ".xlsx",
        ".xls",
    }:

        df = load_excel(
            path,
            sheet_name,
            DEFAULT_MAX_ROWS,
        )


    elif extension == ".json":

        df = load_json(
            path,
            DEFAULT_MAX_ROWS,
        )


    elif extension == ".parquet":

        df = load_parquet(
            path,
            DEFAULT_MAX_ROWS,
        )


    else:

        raise ValueError(
            f"Unsupported format: {extension}"
        )


    metadata[
        "rows_loaded"
    ] = len(df)

    metadata[
        "columns"
    ] = len(df.columns)


    return (
        df,
        metadata,
    )


# ============================================================
# COLUMN ROLE DETECTION
# ============================================================

ROLE_ALIASES = {

    "date": [
        "date",
        "order date",
        "sales date",
        "transaction date",
        "invoice date",
        "created at",
        "created date",
        "timestamp",
        "datetime",
        "month",
        "year",
    ],

    "revenue": [
        "revenue",
        "sales",
        "sale",
        "net sales",
        "gross sales",
        "turnover",
        "sales amount",
        "revenue amount",
        "total sales",
        "amount",
    ],

    "profit": [
        "profit",
        "net profit",
        "gross profit",
        "profit amount",
        "profit value",
        "earnings",
    ],

    "cost": [
        "cost",
        "total cost",
        "cost amount",
        "cogs",
        "cost of goods",
        "cost of goods sold",
    ],

    "quantity": [
        "quantity",
        "qty",
        "units",
        "unit count",
        "volume",
    ],

    "product": [
        "product",
        "product name",
        "item",
        "item name",
        "sku",
        "product id",
        "product code",
    ],

    "region": [
        "region",
        "area",
        "territory",
        "zone",
        "state",
        "province",
        "country",
        "market",
        "location",
    ],

    "category": [
        "category",
        "product category",
        "segment",
        "department",
        "type",
    ],

    "customer": [
        "customer",
        "customer name",
        "client",
        "client name",
        "buyer",
        "account",
        "account name",
    ],
}


def detect_roles(df):

    roles = {}

    columns = {
        str(column):
            normalize(column)
        for column in df.columns
    }


    for role, aliases in ROLE_ALIASES.items():

        found = None


        # Exact.
        for column, normalized in columns.items():

            if any(
                normalized
                == normalize(alias)
                for alias in aliases
            ):

                found = column
                break


        # Partial.
        if found is None:

            for column, normalized in columns.items():

                if any(
                    normalize(alias)
                    in normalized
                    for alias in aliases
                ):

                    found = column
                    break


        roles[
            role
        ] = found


    return roles


# ============================================================
# COLUMN TYPE
# ============================================================

def detect_column_type(series):

    dtype = series.dtype


    if pd.api.types.is_bool_dtype(
        dtype
    ):

        return "boolean"


    if pd.api.types.is_numeric_dtype(
        dtype
    ):

        return "numeric"


    if pd.api.types.is_datetime64_any_dtype(
        dtype
    ):

        return "datetime"


    if series.dtype == "object":

        sample = (
            series
            .dropna()
            .head(100)
        )


        if len(sample):

            try:

                with warnings.catch_warnings():

                    warnings.simplefilter(
                        "ignore"
                    )

                    parsed = pd.to_datetime(
                        sample,
                        errors="coerce",
                    )


                if (
                    parsed.notna().mean()
                    >= 0.8
                ):

                    return "datetime"

            except Exception:
                pass


    non_null = series.notna().sum()

    if non_null == 0:

        return "empty"


    unique = series.nunique(
        dropna=True
    )


    if (
        unique
        / max(non_null, 1)
        <= 0.05
    ):

        return "categorical"


    return "text"


# ============================================================
# COLUMN PROFILE
# ============================================================

def profile_columns(df):

    result = []


    for column in df.columns:

        series = df[
            column
        ]

        missing = int(
            series.isna().sum()
        )

        unique = int(
            series.nunique(
                dropna=True
            )
        )


        item = {

            "column":
                str(column),

            "type":
                detect_column_type(
                    series
                ),

            "dtype":
                str(series.dtype),

            "missing":
                missing,

            "missing_percent":
                round(
                    (
                        missing
                        / max(
                            len(series),
                            1,
                        )
                    )
                    * 100,
                    2,
                ),

            "unique":
                unique,

        }


        if pd.api.types.is_numeric_dtype(
            series
        ):

            values = pd.to_numeric(
                series,
                errors="coerce",
            ).dropna()


            if not values.empty:

                item.update({

                    "min":
                        json_safe(
                            values.min()
                        ),

                    "max":
                        json_safe(
                            values.max()
                        ),

                    "mean":
                        json_safe(
                            values.mean()
                        ),

                    "median":
                        json_safe(
                            values.median()
                        ),

                    "std":
                        json_safe(
                            values.std()
                        ),

                })


        else:

            try:

                item[
                    "top_values"
                ] = (
                    series
                    .dropna()
                    .astype(str)
                    .value_counts()
                    .head(5)
                    .to_dict()
                )

            except Exception:
                pass


        result.append(
            item
        )


    return result


# ============================================================
# QUALITY
# ============================================================

def missing_analysis(df):

    result = {}


    for column in df.columns:

        count = int(
            df[
                column
            ].isna().sum()
        )


        if count:

            result[
                str(column)
            ] = {

                "missing":
                    count,

                "percent":
                    round(
                        (
                            count
                            / max(
                                len(df),
                                1,
                            )
                        )
                        * 100,
                        2,
                    ),
            }


    return result


def duplicate_analysis(df):

    count = int(
        df.duplicated().sum()
    )


    return {

        "duplicate_rows":
            count,

        "duplicate_percent":
            round(
                count
                / max(len(df), 1)
                * 100,
                2,
            ),

    }


def quality_score(df):

    if df.empty:

        return {
            "score": 0,
            "rating": "EMPTY",
        }


    total_cells = (
        df.shape[0]
        * df.shape[1]
    )

    missing_cells = int(
        df.isna().sum().sum()
    )

    duplicate_rows = int(
        df.duplicated().sum()
    )


    score = 100

    score -= (
        missing_cells
        / max(total_cells, 1)
        * 50
    )

    score -= (
        duplicate_rows
        / max(len(df), 1)
        * 30
    )


    score = max(
        0,
        min(
            100,
            score,
        ),
    )


    if score >= 90:
        rating = "EXCELLENT"
    elif score >= 75:
        rating = "GOOD"
    elif score >= 50:
        rating = "FAIR"
    else:
        rating = "NEEDS CLEANING"


    return {

        "score":
            round(score, 1),

        "rating":
            rating,

        "missing_cells":
            missing_cells,

        "duplicate_rows":
            duplicate_rows,

    }


# ============================================================
# NUMERIC STATISTICS
# ============================================================

def numeric_summary(df):

    result = {}


    for column in (
        df.select_dtypes(
            include="number"
        ).columns
    ):

        values = pd.to_numeric(
            df[
                column
            ],
            errors="coerce",
        ).dropna()


        if values.empty:
            continue


        result[
            str(column)
        ] = {

            "count":
                int(values.count()),

            "sum":
                json_safe(
                    values.sum()
                ),

            "mean":
                json_safe(
                    values.mean()
                ),

            "median":
                json_safe(
                    values.median()
                ),

            "min":
                json_safe(
                    values.min()
                ),

            "max":
                json_safe(
                    values.max()
                ),

            "std":
                json_safe(
                    values.std()
                ),

        }


    return result


# ============================================================
# OUTLIERS
# ============================================================

def outlier_analysis(df):

    result = {}


    for column in (
        df.select_dtypes(
            include="number"
        ).columns
    ):

        values = pd.to_numeric(
            df[
                column
            ],
            errors="coerce",
        ).dropna()


        if len(values) < 5:
            continue


        q1 = values.quantile(
            0.25
        )

        q3 = values.quantile(
            0.75
        )

        iqr = q3 - q1


        if iqr == 0:
            continue


        lower = q1 - 1.5 * iqr
        upper = q3 + 1.5 * iqr


        count = int(
            (
                (values < lower)
                |
                (values > upper)
            ).sum()
        )


        if count:

            result[
                str(column)
            ] = {

                "outliers":
                    count,

                "percent":
                    round(
                        count
                        / len(values)
                        * 100,
                        2,
                    ),

                "lower_bound":
                    json_safe(
                        lower
                    ),

                "upper_bound":
                    json_safe(
                        upper
                    ),

            }


    return result


# ============================================================
# CORRELATIONS
# ============================================================

def correlations(df):

    numeric = df.select_dtypes(
        include="number"
    )


    if numeric.shape[1] < 2:
        return []


    matrix = numeric.corr()

    columns = list(
        matrix.columns
    )

    pairs = []


    for i in range(
        len(columns)
    ):

        for j in range(
            i + 1,
            len(columns),
        ):

            value = matrix.loc[
                columns[i],
                columns[j],
            ]


            if pd.isna(value):
                continue


            value = float(value)


            pairs.append({

                "column_1":
                    str(columns[i]),

                "column_2":
                    str(columns[j]),

                "correlation":
                    round(
                        value,
                        4,
                    ),

                "absolute":
                    round(
                        abs(value),
                        4,
                    ),

            })


    pairs.sort(
        key=lambda item:
            item["absolute"],
        reverse=True,
    )


    return pairs[:20]


# ============================================================
# TREND ANALYSIS
# ============================================================

def trend_analysis(
    df,
    roles,
):

    date_column = roles.get(
        "date"
    )

    if not date_column:
        return {}


    dates = pd.to_datetime(
        df[
            date_column
        ],
        errors="coerce",
    )


    mask = dates.notna()


    if not mask.any():
        return {}


    working = df.loc[
        mask
    ].copy()


    working[
        "__date__"
    ] = dates.loc[
        mask
    ]


    metric = roles.get(
        "revenue"
    )


    if (
        not metric
        or metric not in working.columns
        or not pd.api.types.is_numeric_dtype(
            working[
                metric
            ]
        )
    ):

        numeric_columns = list(
            working.select_dtypes(
                include="number"
            ).columns
        )

        metric = (
            numeric_columns[0]
            if numeric_columns
            else None
        )


    if not metric:

        grouped = (
            working
            .assign(
                period=
                    working[
                        "__date__"
                    ]
                    .dt
                    .to_period(
                        "M"
                    )
                    .astype(str)
            )
            .groupby(
                "period"
            )
            .size()
            .reset_index(
                name="count"
            )
        )


        return {

            "date_column":
                date_column,

            "metric":
                "count",

            "data":
                grouped.to_dict(
                    orient="records"
                ),

        }


    grouped = (
        working
        .assign(
            period=
                working[
                    "__date__"
                ]
                .dt
                .to_period(
                    "M"
                )
                .astype(str)
        )
        .groupby(
            "period"
        )[metric]
        .sum()
        .reset_index()
    )


    grouped[
        "value"
    ] = grouped[
        metric
    ]


    grouped[
        "growth_percent"
    ] = (
        grouped[
            "value"
        ]
        .pct_change()
        .replace(
            [
                float("inf"),
                float("-inf"),
            ],
            pd.NA,
        )
        * 100
    )


    return {

        "date_column":
            date_column,

        "metric":
            metric,

        "data":
            grouped[
                [
                    "period",
                    "value",
                    "growth_percent",
                ]
            ].to_dict(
                orient="records"
            ),

    }


# ============================================================
# TOP PRODUCTS
# ============================================================

def top_products(
    df,
    roles,
):

    product = roles.get(
        "product"
    )

    if not product:
        return []


    revenue = roles.get(
        "revenue"
    )


    if (
        revenue
        and revenue in df.columns
        and pd.api.types.is_numeric_dtype(
            df[
                revenue
            ]
        )
    ):

        result = (
            df.groupby(
                product,
                dropna=False,
            )[
                revenue
            ]
            .sum()
            .sort_values(
                ascending=False
            )
            .head(10)
            .reset_index()
        )


        result.columns = [
            "product",
            "value",
        ]


    else:

        result = (
            df[
                product
            ]
            .value_counts()
            .head(10)
            .reset_index()
        )


        result.columns = [
            "product",
            "count",
        ]


    return result.to_dict(
        orient="records"
    )


# ============================================================
# REGIONAL PERFORMANCE
# ============================================================

def regional_performance(
    df,
    roles,
):

    region = roles.get(
        "region"
    )

    if not region:
        return []


    revenue = roles.get(
        "revenue"
    )


    if (
        revenue
        and revenue in df.columns
        and pd.api.types.is_numeric_dtype(
            df[
                revenue
            ]
        )
    ):

        result = (
            df.groupby(
                region,
                dropna=False,
            )[
                revenue
            ]
            .sum()
            .sort_values(
                ascending=False
            )
            .head(20)
            .reset_index()
        )


        result.columns = [
            "region",
            "value",
        ]


    else:

        result = (
            df[
                region
            ]
            .value_counts()
            .head(20)
            .reset_index()
        )


        result.columns = [
            "region",
            "count",
        ]


    return result.to_dict(
        orient="records"
    )


# ============================================================
# PROFIT
# ============================================================

def profit_analysis(
    df,
    roles,
):

    profit_column = roles.get(
        "profit"
    )

    revenue_column = roles.get(
        "revenue"
    )

    cost_column = roles.get(
        "cost"
    )


    derived = False


    if (
        profit_column
        and profit_column in df.columns
    ):

        values = pd.to_numeric(
            df[
                profit_column
            ],
            errors="coerce",
        )

    elif (
        revenue_column
        and cost_column
        and revenue_column in df.columns
        and cost_column in df.columns
    ):

        revenue = pd.to_numeric(
            df[
                revenue_column
            ],
            errors="coerce",
        )

        cost = pd.to_numeric(
            df[
                cost_column
            ],
            errors="coerce",
        )

        values = (
            revenue
            - cost
        )

        derived = True

    else:

        return {}


    values = values.dropna()


    if values.empty:

        return {}


    total_profit = float(
        values.sum()
    )


    result = {

        "derived_profit":
            derived,

        "profit_column":
            profit_column,

        "total_profit":
            json_safe(
                total_profit
            ),

        "average_profit":
            json_safe(
                values.mean()
            ),

        "minimum_profit":
            json_safe(
                values.min()
            ),

        "maximum_profit":
            json_safe(
                values.max()
            ),

    }


    if revenue_column:

        revenue = pd.to_numeric(
            df[
                revenue_column
            ],
            errors="coerce",
        )

        total_revenue = revenue.sum()


        if total_revenue:

            result[
                "profit_margin_percent"
            ] = round(
                total_profit
                / total_revenue
                * 100,
                2,
            )


    return result


# ============================================================
# CATEGORICAL ANALYSIS
# ============================================================

def categorical_analysis(df):

    result = {}


    for column in df.columns:

        if (
            detect_column_type(
                df[
                    column
                ]
            )
            != "categorical"
        ):
            continue


        values = (
            df[
                column
            ]
            .astype(str)
            .value_counts()
            .head(10)
            .reset_index()
        )


        values.columns = [
            "value",
            "count",
        ]


        result[
            str(column)
        ] = values.to_dict(
            orient="records"
        )


    return result


# ============================================================
# INSIGHTS
# ============================================================

def generate_insights(
    df,
    analysis,
):

    insights = []


    insights.append(
        f"The dataset contains "
        f"{len(df):,} rows and "
        f"{len(df.columns):,} columns."
    )


    duplicates = analysis[
        "duplicates"
    ]["duplicate_rows"]


    if duplicates:

        insights.append(
            f"{duplicates:,} duplicate "
            f"rows were detected."
        )

    else:

        insights.append(
            "No duplicate rows were detected."
        )


    missing = sum(
        item.get(
            "missing",
            0,
        )
        for item in analysis[
            "missing_values"
        ].values()
    )


    if missing:

        insights.append(
            f"{missing:,} missing values "
            f"were detected."
        )

    else:

        insights.append(
            "No missing values were detected."
        )


    quality = analysis[
        "quality"
    ]


    insights.append(
        f"Data quality score: "
        f"{quality['score']}/100 "
        f"({quality['rating']})."
    )


    roles = analysis[
        "roles"
    ]


    revenue_column = roles.get(
        "revenue"
    )


    if revenue_column:

        try:

            values = pd.to_numeric(
                df[
                    revenue_column
                ],
                errors="coerce",
            ).dropna()


            if not values.empty:

                insights.append(
                    f"Total "
                    f"{revenue_column}: "
                    f"{values.sum():,.2f}."
                )

        except Exception:
            pass


    profit = analysis[
        "profit"
    ]


    if profit.get(
        "total_profit"
    ) is not None:

        insights.append(
            f"Total profit: "
            f"{profit['total_profit']:,.2f}."
        )


    if profit.get(
        "profit_margin_percent"
    ) is not None:

        insights.append(
            f"Profit margin: "
            f"{profit['profit_margin_percent']:.2f}%."
        )


    products = analysis[
        "top_products"
    ]


    if products:

        first = products[0]

        value = (
            first.get("value")
            if first.get("value") is not None
            else first.get("count")
        )


        insights.append(
            f"Top product: "
            f"{first.get('product')} "
            f"({safe_string(value)})."
        )


    regions = analysis[
        "regional_performance"
    ]


    if regions:

        first = regions[0]

        value = (
            first.get("value")
            if first.get("value") is not None
            else first.get("count")
        )


        insights.append(
            f"Top region: "
            f"{first.get('region')} "
            f"({safe_string(value)})."
        )


    corr = analysis[
        "correlations"
    ]


    if corr:

        strongest = corr[0]

        insights.append(
            f"Strongest numeric correlation: "
            f"{strongest['column_1']} vs "
            f"{strongest['column_2']} "
            f"({strongest['correlation']:.2f})."
        )


    for column, info in list(
        analysis[
            "outliers"
        ].items()
    )[:3]:

        insights.append(
            f"{column}: "
            f"{info['outliers']:,} "
            f"potential outliers."
        )


    return insights


# ============================================================
# COMPLETE ANALYSIS
# ============================================================

def analyze_dataframe(
    df,
    metadata,
):

    roles = detect_roles(
        df
    )


    analysis = {

        "metadata":
            metadata,

        "roles":
            roles,

        "shape": {
            "rows":
                len(df),

            "columns":
                len(df.columns),
        },

        "columns":
            [
                str(column)
                for column in df.columns
            ],

        "column_profiles":
            profile_columns(df),

        "missing_values":
            missing_analysis(df),

        "duplicates":
            duplicate_analysis(df),

        "quality":
            quality_score(df),

        "numeric_summary":
            numeric_summary(df),

        "outliers":
            outlier_analysis(df),

        "correlations":
            correlations(df),

        "categorical":
            categorical_analysis(df),

        "date_trends":
            trend_analysis(
                df,
                roles,
            ),

        "top_products":
            top_products(
                df,
                roles,
            ),

        "regional_performance":
            regional_performance(
                df,
                roles,
            ),

        "profit":
            profit_analysis(
                df,
                roles,
            ),

    }


    analysis[
        "insights"
    ] = generate_insights(
        df,
        analysis,
    )


    return analysis


# ============================================================
# EXCEL REPORT
# ============================================================

def create_excel_report(
    df,
    analysis,
    path,
):

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )


    with pd.ExcelWriter(
        path,
        engine="openpyxl",
    ) as writer:


        # Summary
        pd.DataFrame({

            "Metric": [
                "File",
                "Rows",
                "Columns",
                "Quality Score",
                "Quality Rating",
                "Duplicate Rows",
                "Missing Cells",
            ],

            "Value": [
                analysis[
                    "metadata"
                ].get(
                    "name",
                    "",
                ),

                analysis[
                    "shape"
                ][
                    "rows"
                ],

                analysis[
                    "shape"
                ][
                    "columns"
                ],

                analysis[
                    "quality"
                ][
                    "score"
                ],

                analysis[
                    "quality"
                ][
                    "rating"
                ],

                analysis[
                    "duplicates"
                ][
                    "duplicate_rows"
                ],

                sum(
                    item.get(
                        "missing",
                        0,
                    )
                    for item
                    in analysis[
                        "missing_values"
                    ].values()
                ),

            ],
        }).to_excel(
            writer,
            sheet_name="Executive Summary",
            index=False,
        )


        # Roles
        pd.DataFrame({

            "Role":
                list(
                    analysis[
                        "roles"
                    ].keys()
                ),

            "Column":
                list(
                    analysis[
                        "roles"
                    ].values()
                ),

        }).to_excel(
            writer,
            sheet_name="Detected Roles",
            index=False,
        )


        # Profiles
        pd.DataFrame(
            analysis[
                "column_profiles"
            ]
        ).to_excel(
            writer,
            sheet_name="Column Profile",
            index=False,
        )


        # Numeric
        if analysis[
            "numeric_summary"
        ]:

            numeric_df = (
                pd.DataFrame(
                    analysis[
                        "numeric_summary"
                    ]
                )
                .T
                .reset_index()
                .rename(
                    columns={
                        "index":
                            "Column"
                    }
                )
            )

            numeric_df.to_excel(
                writer,
                sheet_name="Statistics",
                index=False,
            )


        # Missing
        if analysis[
            "missing_values"
        ]:

            pd.DataFrame([
                {
                    "Column":
                        column,

                    "Missing":
                        values[
                            "missing"
                        ],

                    "Percent":
                        values[
                            "percent"
                        ],
                }

                for column, values
                in analysis[
                    "missing_values"
                ].items()
            ]).to_excel(
                writer,
                sheet_name="Missing Values",
                index=False,
            )


        # Products
        if analysis[
            "top_products"
        ]:

            pd.DataFrame(
                analysis[
                    "top_products"
                ]
            ).to_excel(
                writer,
                sheet_name="Top Products",
                index=False,
            )


        # Regions
        if analysis[
            "regional_performance"
        ]:

            pd.DataFrame(
                analysis[
                    "regional_performance"
                ]
            ).to_excel(
                writer,
                sheet_name="Regional Performance",
                index=False,
            )


        # Profit
        if analysis[
            "profit"
        ]:

            pd.DataFrame([
                analysis[
                    "profit"
                ]
            ]).to_excel(
                writer,
                sheet_name="Profit Analysis",
                index=False,
            )


        # Trends
        if analysis[
            "date_trends"
        ].get(
            "data"
        ):

            pd.DataFrame(
                analysis[
                    "date_trends"
                ][
                    "data"
                ]
            ).to_excel(
                writer,
                sheet_name="Trends",
                index=False,
            )


        # Outliers
        if analysis[
            "outliers"
        ]:

            pd.DataFrame([
                {
                    "Column":
                        column,

                    **info,
                }

                for column, info
                in analysis[
                    "outliers"
                ].items()
            ]).to_excel(
                writer,
                sheet_name="Outliers",
                index=False,
            )


        # Correlation
        if analysis[
            "correlations"
        ]:

            pd.DataFrame(
                analysis[
                    "correlations"
                ]
            ).to_excel(
                writer,
                sheet_name="Correlations",
                index=False,
            )


        # Insights
        pd.DataFrame({

            "Insight":
                analysis[
                    "insights"
                ],

        }).to_excel(
            writer,
            sheet_name="Insights",
            index=False,
        )


        # Data preview
        df.head(
            MAX_REPORT_ROWS
        ).to_excel(
            writer,
            sheet_name="Data Preview",
            index=False,
        )


    # Formatting
    if OPENPYXL_AVAILABLE:

        try:

            workbook = (
                openpyxl.load_workbook(
                    path
                )
            )

            for sheet in workbook.worksheets:

                sheet.freeze_panes = "A2"

                for cell in sheet[1]:

                    cell.font = (
                        openpyxl.styles.Font(
                            bold=True
                        )
                    )

                for column_cells in sheet.columns:

                    try:

                        letter = (
                            column_cells[
                                0
                            ].column_letter
                        )

                    except Exception:

                        continue

                    width = 0

                    for cell in column_cells[:100]:

                        width = max(
                            width,
                            len(
                                safe_string(
                                    cell.value
                                )
                            ),
                        )

                    sheet.column_dimensions[
                        letter
                    ].width = min(
                        width + 2,
                        45,
                    )


            workbook.save(
                path
            )

        except Exception as e:

            debug(
                f"Excel formatting warning: {e}"
            )


    return path


# ============================================================
# HTML DASHBOARD
# ============================================================

def create_html_dashboard(
    df,
    analysis,
    path,
):

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )


    charts = []


    if PLOTLY_AVAILABLE:

        try:

            trends = analysis[
                "date_trends"
            ]


            if trends.get(
                "data"
            ):

                data = pd.DataFrame(
                    trends[
                        "data"
                    ]
                )


                if (
                    "period" in data.columns
                    and
                    "value" in data.columns
                ):

                    charts.append(
                        px.line(
                            data,
                            x="period",
                            y="value",
                            markers=True,
                            title="Trend Over Time",
                        ).to_html(
                            full_html=False,
                            include_plotlyjs="cdn",
                        )
                    )


        except Exception:
            pass


        try:

            products = analysis[
                "top_products"
            ]


            if products:

                data = pd.DataFrame(
                    products
                )


                value_column = (
                    "value"
                    if "value"
                    in data.columns
                    else
                    "count"
                )


                if (
                    "product"
                    in data.columns
                ):

                    charts.append(
                        px.bar(
                            data.sort_values(
                                value_column
                            ),
                            x=value_column,
                            y="product",
                            orientation="h",
                            title="Top Products",
                        ).to_html(
                            full_html=False,
                            include_plotlyjs=False,
                        )
                    )


        except Exception:
            pass


        try:

            regions = analysis[
                "regional_performance"
            ]


            if regions:

                data = pd.DataFrame(
                    regions
                )


                value_column = (
                    "value"
                    if "value"
                    in data.columns
                    else
                    "count"
                )


                if (
                    "region"
                    in data.columns
                ):

                    charts.append(
                        px.bar(
                            data,
                            x="region",
                            y=value_column,
                            title="Regional Performance",
                        ).to_html(
                            full_html=False,
                            include_plotlyjs=False,
                        )
                    )


        except Exception:
            pass


    charts_html = "\n".join(
        charts
    )


    if not charts_html:

        charts_html = (
            "<p>"
            "Interactive charts are unavailable "
            "because Plotly is not installed. "
            "The dashboard still contains the "
            "complete analysis tables and KPIs."
            "</p>"
        )


    roles_html = ""


    for role, column in analysis[
        "roles"
    ].items():

        roles_html += (
            "<tr>"
            f"<td>{safe_string(role)}</td>"
            f"<td>{safe_string(column)}</td>"
            "</tr>"
        )


    profile_html = ""


    for item in analysis[
        "column_profiles"
    ]:

        profile_html += (
            "<tr>"
            f"<td>{safe_string(item.get('column'))}</td>"
            f"<td>{safe_string(item.get('type'))}</td>"
            f"<td>{item.get('missing', 0):,}</td>"
            f"<td>{item.get('missing_percent', 0)}%</td>"
            f"<td>{item.get('unique', 0):,}</td>"
            "</tr>"
        )


    insights_html = ""


    for insight in analysis[
        "insights"
    ]:

        insights_html += (
            f"<div class='insight'>{safe_string(insight)}</div>"
        )


    quality = analysis[
        "quality"
    ]


    profit = analysis[
        "profit"
    ]


    missing_count = sum(
        item.get(
            "missing",
            0,
        )
        for item
        in analysis[
            "missing_values"
        ].values()
    )


    html = f"""
<!DOCTYPE html>
<html lang="en">

<head>

<meta charset="UTF-8">

<meta name="viewport"
      content="width=device-width,
               initial-scale=1.0">

<title>
JARVIS Data Intelligence Dashboard
</title>

<style>

body {{
    margin: 0;
    font-family: Arial, Helvetica, sans-serif;
    background: #f4f6f8;
    color: #202124;
}}

.container {{
    max-width: 1500px;
    margin: auto;
    padding: 30px;
}}

.header {{
    background: #111827;
    color: white;
    padding: 30px;
    border-radius: 18px;
    margin-bottom: 25px;
}}

.cards {{
    display: grid;
    grid-template-columns:
        repeat(auto-fit, minmax(190px, 1fr));
    gap: 18px;
    margin-bottom: 25px;
}}

.card {{
    background: white;
    padding: 22px;
    border-radius: 15px;
    box-shadow: 0 3px 12px rgba(0,0,0,.08);
}}

.label {{
    color: #6b7280;
    font-size: 13px;
}}

.value {{
    font-size: 27px;
    font-weight: bold;
    margin-top: 8px;
}}

.section {{
    background: white;
    padding: 25px;
    border-radius: 15px;
    margin-bottom: 25px;
    box-shadow: 0 3px 12px rgba(0,0,0,.08);
}}

table {{
    width: 100%;
    border-collapse: collapse;
}}

th, td {{
    padding: 10px;
    text-align: left;
    border-bottom: 1px solid #e5e7eb;
}}

th {{
    background: #f9fafb;
}}

.insight {{
    background: #f9fafb;
    padding: 13px;
    margin: 8px 0;
    border-left: 4px solid #374151;
    border-radius: 6px;
}}

.footer {{
    text-align: center;
    color: #6b7280;
    padding: 25px;
}}

</style>

</head>

<body>

<div class="container">

<div class="header">

<h1>
JARVIS Data Intelligence Dashboard
</h1>

<p>
<strong>Dataset:</strong>
{safe_string(analysis["metadata"].get("name"))}
</p>

<p>
<strong>Type:</strong>
{safe_string(analysis["metadata"].get("type"))}
</p>

</div>


<div class="cards">

<div class="card">
<div class="label">Rows</div>
<div class="value">
{analysis["shape"]["rows"]:,}
</div>
</div>

<div class="card">
<div class="label">Columns</div>
<div class="value">
{analysis["shape"]["columns"]:,}
</div>
</div>

<div class="card">
<div class="label">Quality</div>
<div class="value">
{quality["score"]}/100
</div>
</div>

<div class="card">
<div class="label">Duplicates</div>
<div class="value">
{analysis["duplicates"]["duplicate_rows"]:,}
</div>
</div>

<div class="card">
<div class="label">Missing</div>
<div class="value">
{missing_count:,}
</div>
</div>

<div class="card">
<div class="label">Profit</div>
<div class="value">
{safe_string(profit.get("total_profit", "N/A"))}
</div>
</div>

</div>


<div class="section">

<h2>Key Insights</h2>

{insights_html}

</div>


<div class="section">

<h2>Detected Data Roles</h2>

<table>

<tr>
<th>Role</th>
<th>Column</th>
</tr>

{roles_html}

</table>

</div>


<div class="section">

<h2>Charts</h2>

{charts_html}

</div>


<div class="section">

<h2>Column Profile</h2>

<table>

<tr>
<th>Column</th>
<th>Type</th>
<th>Missing</th>
<th>Missing %</th>
<th>Unique</th>
</tr>

{profile_html}

</table>

</div>


<div class="footer">

Generated by JARVIS Data Agent

</div>

</div>

</body>

</html>
"""


    path.write_text(
        html,
        encoding="utf-8",
    )


    return path


# ============================================================
# REPORT GENERATOR
# ============================================================

def generate_reports(
    df,
    analysis,
    dataset_path,
    request,
):

    formats = requested_formats(
        request
    )


    # Plain "analyze" also creates the normal Excel report.
    make_excel = (
        formats["excel"]
        or
        "analyze" in normalize(request)
        or
        "analyse" in normalize(request)
    )


    make_html = (
        formats["html"]
        or
        formats["dashboard"]
    )


    output_dir = ensure_output_dir()


    safe_name = re.sub(
        r"[^A-Za-z0-9_-]+",
        "_",
        dataset_path.stem,
    )


    stamp = timestamp()


    reports = {}


    if make_excel:

        excel_path = (
            output_dir
            /
            (
                f"{safe_name}_"
                f"JARVIS_Report_"
                f"{stamp}.xlsx"
            )
        )


        try:

            create_excel_report(
                df,
                analysis,
                excel_path,
            )

            reports[
                "excel_report"
            ] = str(
                excel_path
            )

        except Exception as e:

            debug(
                f"Excel report failed: {e}"
            )


    if make_html:

        html_path = (
            output_dir
            /
            (
                f"{safe_name}_"
                f"JARVIS_Dashboard_"
                f"{stamp}.html"
            )
        )


        try:

            create_html_dashboard(
                df,
                analysis,
                html_path,
            )

            reports[
                "html_dashboard"
            ] = str(
                html_path
            )

        except Exception as e:

            debug(
                f"HTML dashboard failed: {e}"
            )


    return reports


# ============================================================
# RESULT FORMAT
# ============================================================

def format_result(
    analysis,
    reports,
):

    lines = []


    lines.append(
        "DATASET ANALYSIS"
    )

    lines.append(
        "--------------------------------------------------"
    )


    metadata = analysis[
        "metadata"
    ]


    lines.append(
        f"File: "
        f"{metadata.get('name', 'Unknown')}"
    )


    lines.append(
        f"Type: "
        f"{metadata.get('type', 'Unknown')}"
    )


    lines.append(
        f"Rows: "
        f"{analysis['shape']['rows']:,}"
    )


    lines.append(
        f"Columns: "
        f"{analysis['shape']['columns']:,}"
    )


    lines.append(
        f"Data Quality: "
        f"{analysis['quality']['score']}/100 "
        f"({analysis['quality']['rating']})"
    )


    lines.append(
        f"Duplicate Rows: "
        f"{analysis['duplicates']['duplicate_rows']:,}"
    )


    lines.append("")


    lines.append(
        "DETECTED ROLES"
    )


    for role, column in analysis[
        "roles"
    ].items():

        if column:

            lines.append(
                f"- {role}: {column}"
            )


    lines.append("")


    lines.append(
        "KEY INSIGHTS"
    )


    for index, insight in enumerate(
        analysis[
            "insights"
        ],
        1,
    ):

        lines.append(
            f"{index}. {insight}"
        )


    if analysis[
        "profit"
    ]:

        lines.append("")

        lines.append(
            "PROFIT ANALYSIS"
        )


        if analysis[
            "profit"
        ].get(
            "total_profit"
        ) is not None:

            lines.append(
                "Total Profit: "
                f"{analysis['profit']['total_profit']:,.2f}"
            )


        if analysis[
            "profit"
        ].get(
            "profit_margin_percent"
        ) is not None:

            lines.append(
                "Profit Margin: "
                f"{analysis['profit']['profit_margin_percent']:.2f}%"
            )


    if analysis[
        "top_products"
    ]:

        lines.append("")

        lines.append(
            "TOP PRODUCTS"
        )


        for index, item in enumerate(
            analysis[
                "top_products"
            ][:10],
            1,
        ):

            value = (
                item.get("value")
                if item.get("value") is not None
                else item.get("count")
            )


            lines.append(
                f"{index}. "
                f"{item.get('product')} "
                f"({safe_string(value)})"
            )


    if analysis[
        "regional_performance"
    ]:

        lines.append("")

        lines.append(
            "REGIONAL PERFORMANCE"
        )


        for index, item in enumerate(
            analysis[
                "regional_performance"
            ][:10],
            1,
        ):

            value = (
                item.get("value")
                if item.get("value") is not None
                else item.get("count")
            )


            lines.append(
                f"{index}. "
                f"{item.get('region')} "
                f"({safe_string(value)})"
            )


    if analysis[
        "outliers"
    ]:

        lines.append("")

        lines.append(
            "OUTLIERS"
        )


        for column, info in list(
            analysis[
                "outliers"
            ].items()
        )[:10]:

            lines.append(
                f"- {column}: "
                f"{info['outliers']:,}"
            )


    if analysis[
        "correlations"
    ]:

        lines.append("")

        lines.append(
            "STRONGEST CORRELATIONS"
        )


        for item in analysis[
            "correlations"
        ][:5]:

            lines.append(
                f"- "
                f"{item['column_1']} vs "
                f"{item['column_2']}: "
                f"{item['correlation']:.2f}"
            )


    if reports:

        lines.append("")

        lines.append(
            "GENERATED REPORTS"
        )

        lines.append(
            "--------------------------------------------------"
        )


        if reports.get(
            "excel_report"
        ):

            lines.append(
                "Excel Report: "
                +
                reports[
                    "excel_report"
                ]
            )


        if reports.get(
            "html_dashboard"
        ):

            lines.append(
                "HTML Dashboard: "
                +
                reports[
                    "html_dashboard"
                ]
            )


    return "\n".join(
        lines
    )


# ============================================================
# DATASET SELECTION
# ============================================================

def choose_dataset(
    datasets,
    request,
):

    value = normalize(
        request
    )


    # Filename match.
    for path in datasets:

        if normalize(
            path.name
        ) in value:

            return path


    # Stem match.
    for path in datasets:

        if normalize(
            path.stem
        ) in value:

            return path


    # One dataset = automatic.
    if len(
        datasets
    ) == 1:

        return datasets[0]


    return None


def selection_message(
    folder,
    datasets,
):

    lines = [

        f"I found {len(datasets)} supported datasets.",

        f"Folder: {folder}",

        "",

        "Choose one:",

        "",

    ]


    for index, path in enumerate(
        datasets[
            :MAX_FILES_TO_DISPLAY
        ],
        1,
    ):

        try:

            size = human_size(
                path.stat().st_size
            )

        except Exception:

            size = "unknown size"


        relative = str(
            path.relative_to(
                folder
            )
        )


        lines.append(
            f"{index}. "
            f"{relative} "
            f"({file_type(path)}, {size})"
        )


    if len(datasets) > MAX_FILES_TO_DISPLAY:

        lines.append(
            f"... plus "
            f"{len(datasets) - MAX_FILES_TO_DISPLAY} "
            f"more."
        )


    lines.append("")

    lines.append(
        "Reply with the number or filename."
    )


    return "\n".join(
        lines
    )


def resolve_pending_selection(
    request,
):

    if not _PENDING_DATASETS:

        return None


    value = normalize(
        request
    )


    if value.isdigit():

        index = (
            int(value)
            - 1
        )


        if (
            0 <= index
            < len(
                _PENDING_DATASETS
            )
        ):

            return (
                _PENDING_DATASETS[
                    index
                ]
            )


    for path in _PENDING_DATASETS:

        if normalize(
            path.name
        ) == value:

            return path


    for path in _PENDING_DATASETS:

        if normalize(
            path.stem
        ) == value:

            return path


    return None


# ============================================================
# MAIN DATA FUNCTION
# ============================================================

def data_analyze(
    request,
):

    global _PENDING_DATASETS
    global _PENDING_REQUEST


    request = str(
        request or ""
    ).strip()


    if not request:

        return {
            "success": False,
            "message":
                "Please provide a dataset request.",
        }


    # ========================================================
    # PENDING SELECTION
    # ========================================================

    selected_path = resolve_pending_selection(
        request
    )


    if selected_path:

        original_request = (
            _PENDING_REQUEST
        )

        clear_pending_selection()

        log(
            f"Selected dataset: {selected_path}"
        )


    else:

        original_request = request

        selected_path = None


        # ====================================================
        # PATH
        # ====================================================

        path = extract_windows_path(
            request
        )


        if path is None:

            return {
                "success": False,
                "message": (
                    "I couldn't find a valid dataset "
                    "file or folder path.\n\n"
                    "Example:\n"
                    r"analyze C:\Users\Soura\Downloads"
                ),
            }


        log(
            f"Path: {path}"
        )


        # ====================================================
        # DOES PATH EXIST?
        # ====================================================

        if not path.exists():

            return {
                "success": False,
                "message": (
                    "I couldn't find this path:\n"
                    f"{path}\n\n"
                    "Please give me the correct "
                    "file or folder path."
                ),
            }


        # ====================================================
        # DIRECT FILE
        # ====================================================

        if path.is_file():

            if not is_supported_file(
                path
            ):

                return {
                    "success": False,
                    "message": (
                        f"{path.name} is not a supported "
                        "dataset type.\n\n"
                        "Supported: CSV, XLSX, XLS, JSON, Parquet."
                    ),
                }


            selected_path = path


        # ====================================================
        # FOLDER
        # ====================================================

        elif path.is_dir():

            log(
                "Scanning folder and subfolders..."
            )


            datasets = discover_datasets(
                path,
                recursive=True,
            )


            if not datasets:

                return {
                    "success": False,
                    "message": (
                        f"No supported datasets were found "
                        f"in {path} or its subfolders.\n\n"
                        "Supported: CSV, XLSX, XLS, JSON, Parquet."
                    ),
                }


            selected_path = choose_dataset(
                datasets,
                original_request,
            )


            if selected_path is None:

                _PENDING_DATASETS = datasets

                _PENDING_REQUEST = (
                    original_request
                )


                return {

                    "success":
                        True,

                    "needs_selection":
                        True,

                    "message":
                        selection_message(
                            path,
                            datasets,
                        ),

                }


    # ========================================================
    # EXCEL SHEET
    # ========================================================

    sheet_name = None


    if selected_path.suffix.lower() in {
        ".xlsx",
        ".xls",
    }:

        sheets = get_excel_sheets(
            selected_path
        )


        if len(sheets) > 1:

            request_text = normalize(
                original_request
            )


            for sheet in sheets:

                if normalize(
                    sheet
                ) in request_text:

                    sheet_name = sheet

                    break


            if sheet_name is None:

                sheet_name = sheets[0]

                log(
                    f"Workbook has {len(sheets)} sheets."
                )

                log(
                    f"Using first sheet: {sheet_name}"
                )


    # ========================================================
    # LOAD
    # ========================================================

    log(
        f"Loading: {selected_path}"
    )


    try:

        df, metadata = load_dataset(
            selected_path,
            sheet_name,
        )

    except Exception as e:

        debug(
            f"{type(e).__name__}: {e}"
        )

        return {

            "success":
                False,

            "message":
                (
                    "The dataset was found, "
                    "but I could not read it:\n"
                    f"{e}"
                ),

        }


    if df.empty:

        return {

            "success":
                False,

            "message":
                "The dataset contains no readable rows.",

        }


    log(
        f"Loaded {len(df):,} rows x "
        f"{len(df.columns):,} columns."
    )


    # ========================================================
    # ANALYZE
    # ========================================================

    log(
        "Running full analysis..."
    )


    analysis = analyze_dataframe(
        df,
        metadata,
    )


    # ========================================================
    # REPORTS
    # ========================================================

    log(
        "Generating reports..."
    )


    reports = generate_reports(
        df,
        analysis,
        selected_path,
        original_request,
    )


    # ========================================================
    # RESULT
    # ========================================================

    return {

        "success":
            True,

        "message":
            format_result(
                analysis,
                reports,
            ),

        "dataset":
            str(selected_path),

        "analysis":
            analysis,

        "reports":
            reports,

    }


# ============================================================
# COMPATIBILITY ALIASES
# ============================================================

def analyze(
    request,
):

    return data_analyze(
        request
    )


def research_data(
    request,
):

    return data_analyze(
        request
    )


# ============================================================
# TEST
# ============================================================

if __name__ == "__main__":

    print(
        "=" * 60
    )

    print(
        "JARVIS DATA AGENT"
    )

    print(
        "=" * 60
    )

    print(
        "Supported formats:"
    )

    for extension in sorted(
        SUPPORTED_EXTENSIONS
    ):

        print(
            f"  - {extension}"
        )

    print()

    print(
        "OpenPyXL: "
        +
        (
            "ENABLED"
            if OPENPYXL_AVAILABLE
            else "UNAVAILABLE"
        )
    )

    print(
        "Plotly: "
        +
        (
            "ENABLED"
            if PLOTLY_AVAILABLE
            else "UNAVAILABLE"
        )
    )

    print()

    print(
        "JARVIS Data Agent loaded successfully."
    )